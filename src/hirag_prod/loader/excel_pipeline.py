import logging
from typing import Dict, List, Optional, Tuple
from urllib.parse import unquote, urlparse

import json_repair
from openpyxl import Workbook, load_workbook

from hirag_prod._utils import log_error_info
from hirag_prod.configs.functions import get_config_manager, get_llm_config
from hirag_prod.loader.utils import route_file_path
from hirag_prod.prompt import PROMPTS
from hirag_prod.resources.functions import get_chat_service

logger = logging.getLogger(__name__)

_WORKBOOK_CACHE: Dict[str, "Workbook"] = {}


def _resolve_local_path(excel_uri: str) -> Optional[str]:
    if not excel_uri:
        return None
    try:
        if excel_uri.startswith("file://"):
            p = urlparse(excel_uri)
            raw_path = f"/{p.netloc}{p.path}" if p.netloc else p.path
            return unquote(raw_path)
        try:
            return route_file_path("excel_loader", excel_uri)
        except Exception:
            return excel_uri
    except Exception as e:
        log_error_info(logging.ERROR, "Failed to resolve excel path", e)
        return None


def _get_workbook(excel_uri: str):
    local = _resolve_local_path(excel_uri)
    if not local:
        return None
    if local in _WORKBOOK_CACHE:
        return _WORKBOOK_CACHE[local]
    try:
        wb = load_workbook(local, data_only=True, read_only=True)
        _WORKBOOK_CACHE[local] = wb
        return wb
    except Exception as e:
        log_error_info(logging.ERROR, f"Failed to open workbook: {local}", e)
        return None


def _find_first_cell_containing(ws, needle: str) -> Optional[Tuple[int, int]]:
    if not needle:
        return None
    target = str(needle).strip()
    for row in ws.iter_rows():
        for cell in row:
            try:
                if cell.value is not None and target in str(cell.value):
                    return (cell.row, cell.column)
            except Exception:
                continue
    return None


def _find_cell_by_xy(ws, x_label: str, y_label: str) -> Optional[Tuple[int, int]]:
    x_pos = _find_first_cell_containing(ws, x_label) if x_label else None
    y_pos = _find_first_cell_containing(ws, y_label) if y_label else None

    if x_pos and y_pos:
        row = x_pos[0] if x_pos[1] < y_pos[1] else y_pos[0]
        column = y_pos[1] if x_pos[1] < y_pos[1] else x_pos[1]
        return (row, column)
    else:
        return None


async def _extract_excel_cell_coord_by_llm(
    text_to_be_cited: str, excel_latex: str
) -> Tuple[Optional[str], Optional[str]]:
    try:
        locate_excel_cell_prompt = PROMPTS[
            "locate_excel_cell_" + get_config_manager().language
        ].format(text_to_be_cited=text_to_be_cited, excel_latex=excel_latex)
        excel_cell_coord_result = await get_chat_service().complete(
            prompt=locate_excel_cell_prompt,
            model=get_llm_config().model_name,
            timeout=get_llm_config().timeout,
            max_tokens=get_llm_config().max_tokens,
        )
        decoded_obj = json_repair.repair_json(
            excel_cell_coord_result, return_objects=True
        )
        x = str(decoded_obj.get("x") or "").strip() or None
        y = str(decoded_obj.get("y") or "").strip() or None
        return x, y
    except Exception as e:
        log_error_info(logging.ERROR, "excel cell coordinate extraction failed", e)
        return None, None


async def annotate_excel_cell_bbox(
    text_to_cite: str, chunk_like: Dict
) -> Optional[List[int]]:
    """
    Compute numeric bbox for an excel_sheet chunk:
    - Uses LLM to get x/y labels from the sentence and the chunk's LaTeX.
    - Finds cell via openpyxl (sheet from headers[0], workbook from uri).
    - Returns bbox as [col_index, row_index] (1-based), or None if not found.
    """
    try:
        if not chunk_like:
            return None
        chunk_type = (
            chunk_like.get("chunkType") or chunk_like.get("type") or ""
        ).lower()
        if chunk_type != "excel_sheet":
            return None

        excel_latex = (chunk_like.get("text") or "").strip()
        if not excel_latex:
            return None

        headers = chunk_like.get("headers") or []
        sheet_name = headers[0] if headers else None
        if not sheet_name:
            return None

        excel_uri = chunk_like.get("uri") or ""
        wb = _get_workbook(excel_uri)
        if wb is None or sheet_name not in wb.sheetnames:
            return None

        ws = wb[sheet_name]
        x_label, y_label = await _extract_excel_cell_coord_by_llm(
            text_to_cite, excel_latex
        )
        if not (x_label or y_label):
            return None
        breakpoint()
        pos = _find_cell_by_xy(ws, x_label, y_label)
        if not pos:
            return None

        row_idx, col_idx = pos[0], pos[1]
        return [int(col_idx), int(row_idx)]
    except Exception as e:
        log_error_info(logging.ERROR, "annotate_excel_cell_bbox failed", e)
        return None
